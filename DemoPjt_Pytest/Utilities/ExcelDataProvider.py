#E:\VidyaWorkspace\PythonWorkspace\PythonProjects\SeleniumPythonBasics\DemoPjt_Pytest\ExcelFiles\DemoPjt_TestData.xlsx
import openpyxl
import os

path = os.path.join(os.path.abspath(os.pardir), 'ExcelFiles\DemoPjt_TestData.xlsx')
print(path)

def getRowCount(file, sheetname):
    wb = openpyxl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheetname)
    return (sheet.max_row)
#print(getRowCount(path, 'Sheet1'))

def getColumnCount(file, sheetname):
    wb = openpyxl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheetname)
    return (sheet.max_column)
#print(getColumnCount(path,'Sheet1'))

def readData(file, sheetname, rowNum, colNum):
    wb = openpyxl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheetname)
    return sheet.cell(row=rowNum, column=colNum).value
#print(readData(path, 'Sheet1', 1, 2))


def writeData(file, sheetname, rowNum, colNum, data):
    wb = openpyxl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheetname)
    sheet.cell(row=rowNum, column=colNum).value = data
    wb.save(file)
#print(writeData(path, 'Sheet1', 11, 11, 'My Data Written By - Vidya'))