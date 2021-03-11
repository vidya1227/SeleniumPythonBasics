# Reference:
# https://www.geeksforgeeks.org/python-reading-excel-file-using-openpyxl-module/?ref=rp
# 1 : Program to print the particular cell value
# Python program to read an excel file
# import openpyxl module
# E:\SampleTestData.xlsx
import openpyxl

# Give the location of the file
path = "E:\SampleTestData.xlsx"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

# Cell objects also have row, column,
# and coordinate attributes that provide
# location information for the cell.

# Note: The first row or
# column integer is 1, not 0.

# Cell object is created by using
# sheet object's cell() method.
cell_obj = sheet_obj.cell(row=1, column=3)

# Print value of cell object
# using the value attribute
print(cell_obj.value)

# print the total number of rows
print(sheet_obj.max_row)

# ptint total number of column
print(sheet_obj.max_column)

#Print all columns name
max_col = sheet_obj.max_column

# Loop will print all columns name
for i in range(1, max_col + 1):
    cell_obj = sheet_obj.cell(row=1, column=i)
    print(cell_obj.value)

#Print first column value
m_row = sheet_obj.max_row
# Loop will print all values
# of first column
for i in range(1, m_row + 1):
    cell_obj = sheet_obj.cell(row=i, column=1)
    print(cell_obj.value)

#Print a particular row value
max_col1 = sheet_obj.max_column

# Will print a particular row value
for i in range(1, max_col1 + 1):
    cell_obj = sheet_obj.cell(row=2, column=i)
    print(cell_obj.value, end=" ")
