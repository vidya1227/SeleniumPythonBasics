import openpyxl

def getRowCount(file, sheetname):
    wb = openpyxl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheetname)
    return (sheet.max_row)

def getColumnCount(file, sheetname):
    wb = openpyxl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheetname)
    return (sheet.max_column)

def readData(file, sheetname, rowNum, colNum):
    wb = openpyxl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheetname)
    return sheet.cell(row=rowNum, column=colNum).value

def writeData(file, sheetname, rowNum, colNum, data):
    wb = openpyxl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheetname)
    sheet.cell(row=rowNum, column=colNum).value = data
    wb.save(file)

