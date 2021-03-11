import openpyxl


excelPAth = 'E:\SampleTestData.xlsx'

wb = openpyxl.load_workbook(excelPAth)
sheet = wb.active
#wb.get_sheet_by_name('Sheet1')

rows =sheet.max_row
print(rows)
columns = sheet.max_column
print(columns)

for r in range(1, rows+1):
    for c in range(1, columns+1):
        print(sheet.cell(row=r, column=c).value, end="      ")

    print()
